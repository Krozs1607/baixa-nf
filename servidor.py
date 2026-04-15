"""
Servidor Web - Painel de Controle da Automação de Baixa de NF
Roda em http://localhost:5000
"""

import os
import json
import threading
import logging
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Flask, render_template_string, jsonify, request, send_file
from automacao_baixa import AutomacaoBaixa
from automacao_gaulesa import AutomacaoGaulesa
from automacao_cancelamento import AutomacaoCancelamento

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB max upload

# Estado global da automação
estado = {
    "browser_aberto": False,
    "rodando": False,
    "loja_selecionada": None,
    "progresso": {"total": 0, "processadas": 0, "sucesso": 0, "pago": 0, "nao_encontrada": 0, "baixada_anteriormente": 0, "erro": 0},
    "log_mensagens": [],
    "nf_atual": "",
    "tabela_nfs": [],  # Lista de {nf, status, mensagem} para a tabela de monitoramento
    "tabela_analise": [],  # Lista para análise (não faz baixa)
    "dealer_pronto": False,  # True quando o script detectou o frame
    "inicio_confirmado": False,  # True quando o usuário clicou "Iniciar Baixas"
}

automacao: AutomacaoBaixa = None
automacao_gaulesa: AutomacaoGaulesa = None
automacao_cancel: AutomacaoCancelamento = None

# Mapeamento loja -> arquivo Excel
LOJAS = {
    "mandarim_iguatemi": {"nome": "Mandarim Iguatemi", "arquivo": "Mandarim Iguatemi.xlsx"},
    "mandarim_itabuna": {"nome": "Mandarim Itabuna", "arquivo": "Mandarim Itabuna.xlsx"},
    "mandarim_lauro": {"nome": "Mandarim Lauro de Freitas", "arquivo": "Mandarim Lauro de Freitas.xlsx"},
}

HTML_PAGE = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Automacao de Baixa - Dealer.net</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Segoe UI', Tahoma, sans-serif;
            background: #1a1a2e;
            color: #eee;
            min-height: 100vh;
            display: flex;
            flex-direction: column;
            align-items: center;
            padding: 30px 20px;
        }
        h1 {
            font-size: 28px;
            margin-bottom: 8px;
            color: #e94560;
        }
        .subtitle {
            color: #888;
            margin-bottom: 30px;
            font-size: 14px;
        }
        .card {
            background: #16213e;
            border-radius: 12px;
            padding: 30px;
            width: 100%;
            max-width: 600px;
            margin-bottom: 20px;
            border: 1px solid #0f3460;
        }
        .card h2 {
            font-size: 18px;
            margin-bottom: 20px;
            color: #e94560;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        .card h2 .step {
            background: #e94560;
            color: white;
            width: 28px;
            height: 28px;
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 14px;
            flex-shrink: 0;
        }
        select {
            width: 100%;
            padding: 12px 16px;
            border-radius: 8px;
            border: 2px solid #0f3460;
            background: #1a1a2e;
            color: #eee;
            font-size: 16px;
            cursor: pointer;
            margin-bottom: 16px;
        }
        select:focus { border-color: #e94560; outline: none; }
        .file-upload {
            margin-bottom: 16px;
        }
        .file-label {
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 14px 16px;
            border: 2px dashed #0f3460;
            border-radius: 8px;
            background: #1a1a2e;
            color: #888;
            cursor: pointer;
            transition: all 0.2s;
            font-size: 14px;
        }
        .file-label:hover {
            border-color: #e94560;
            color: #e94560;
            background: #16213e;
        }
        .file-label.has-file {
            border-style: solid;
            border-color: #0cca4a;
            color: #0cca4a;
            background: #0cca4a11;
        }
        .btn {
            width: 100%;
            padding: 14px;
            border: none;
            border-radius: 8px;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.2s;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
        }
        .btn:disabled {
            opacity: 0.4;
            cursor: not-allowed;
        }
        .btn-primary {
            background: #e94560;
            color: white;
        }
        .btn-primary:hover:not(:disabled) { background: #c73652; }
        .btn-success {
            background: #0cca4a;
            color: white;
        }
        .btn-success:hover:not(:disabled) { background: #0aa83e; }
        .btn-danger {
            background: #dc3545;
            color: white;
            margin-top: 10px;
        }
        .btn-danger:hover:not(:disabled) { background: #b52d3a; }
        .status-bar {
            display: flex;
            gap: 12px;
            margin-bottom: 16px;
            flex-wrap: wrap;
        }
        .status-item {
            background: #1a1a2e;
            padding: 10px 16px;
            border-radius: 8px;
            text-align: center;
            flex: 1;
            min-width: 80px;
        }
        .status-item .number {
            font-size: 24px;
            font-weight: 700;
        }
        .status-item .label {
            font-size: 11px;
            color: #888;
            margin-top: 2px;
        }
        .status-item.ok .number { color: #0cca4a; }
        .status-item.pago .number { color: #ffc107; }
        .status-item.erro .number { color: #dc3545; }
        .status-item.total .number { color: #17a2b8; }
        .progress-bar-container {
            width: 100%;
            height: 8px;
            background: #1a1a2e;
            border-radius: 4px;
            margin-bottom: 16px;
            overflow: hidden;
        }
        .progress-bar {
            height: 100%;
            background: linear-gradient(90deg, #e94560, #0cca4a);
            border-radius: 4px;
            transition: width 0.5s;
            width: 0%;
        }
        .log-box {
            background: #0d1117;
            border-radius: 8px;
            padding: 16px;
            max-height: 200px;
            overflow-y: auto;
            font-family: 'Consolas', 'Courier New', monospace;
            font-size: 12px;
            line-height: 1.6;
            border: 1px solid #21262d;
        }
        .log-box .log-line { padding: 1px 0; }
        .log-line.success { color: #0cca4a; }
        .log-line.error { color: #dc3545; }
        .log-line.warning { color: #ffc107; }
        .log-line.info { color: #58a6ff; }
        .nf-table-container {
            max-height: 350px;
            overflow-y: auto;
            border-radius: 8px;
            border: 1px solid #21262d;
            margin-bottom: 16px;
        }
        .nf-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
        }
        .nf-table thead { position: sticky; top: 0; z-index: 1; }
        .nf-table th {
            background: #0f3460;
            padding: 10px 12px;
            text-align: left;
            font-weight: 600;
            color: #ccc;
            font-size: 11px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        .nf-table td {
            padding: 8px 12px;
            border-bottom: 1px solid #1a1a2e;
        }
        .nf-table tr { background: #0d1117; }
        .nf-table tr:nth-child(even) { background: #111827; }
        .nf-table tr.processando { background: #1a2744; animation: pulse 1.5s infinite; }
        @keyframes pulse {
            0%, 100% { opacity: 1; }
            50% { opacity: 0.7; }
        }
        .badge {
            display: inline-block;
            padding: 3px 10px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
            text-transform: uppercase;
        }
        .badge-sucesso { background: #0cca4a22; color: #0cca4a; border: 1px solid #0cca4a44; }
        .badge-pago { background: #ffc10722; color: #ffc107; border: 1px solid #ffc10744; }
        .badge-erro { background: #dc354522; color: #dc3545; border: 1px solid #dc354544; }
        .badge-nao_encontrada { background: #6c757d22; color: #adb5bd; border: 1px solid #6c757d44; }
        .badge-baixada_anteriormente { background: #ff850022; color: #ff8500; border: 1px solid #ff850044; }
        .badge-processando { background: #17a2b822; color: #17a2b8; border: 1px solid #17a2b844; }
        .badge-aguardando { background: #1a1a2e; color: #555; border: 1px solid #333; }
        .nf-atual {
            text-align: center;
            padding: 8px;
            color: #58a6ff;
            font-size: 14px;
            margin-bottom: 8px;
        }
        .aviso {
            background: #0f3460;
            border-left: 4px solid #ffc107;
            padding: 12px 16px;
            border-radius: 0 8px 8px 0;
            margin-bottom: 16px;
            font-size: 13px;
            color: #ccc;
        }
        .tabs {
            display: flex;
            gap: 4px;
            width: 100%;
            max-width: 900px;
            margin-bottom: 20px;
            border-bottom: 2px solid #0f3460;
        }
        .tab {
            padding: 12px 24px;
            background: transparent;
            color: #888;
            border: none;
            cursor: pointer;
            font-size: 15px;
            font-weight: 600;
            border-radius: 8px 8px 0 0;
            transition: all 0.2s;
        }
        .tab:hover { color: #ccc; }
        .tab.active {
            background: #16213e;
            color: #e94560;
            border-bottom: 2px solid #e94560;
            margin-bottom: -2px;
        }
        .tab-content { display: none; width: 100%; align-items: center; flex-direction: column; }
        .tab-content.active { display: flex; }
        .chart-container {
            background: #16213e;
            border-radius: 12px;
            padding: 30px;
            width: 100%;
            max-width: 900px;
            margin-bottom: 20px;
            border: 1px solid #0f3460;
        }
        .chart-wrapper {
            max-width: 400px;
            margin: 0 auto;
        }
        .chart-stats {
            display: flex;
            gap: 12px;
            margin-top: 20px;
            flex-wrap: wrap;
            justify-content: center;
        }
        .chart-stat {
            background: #1a1a2e;
            padding: 12px 20px;
            border-radius: 8px;
            text-align: center;
            min-width: 120px;
        }
        .chart-stat .dot {
            display: inline-block;
            width: 10px;
            height: 10px;
            border-radius: 50%;
            margin-right: 6px;
        }
        .chart-stat .number {
            font-size: 22px;
            font-weight: 700;
            color: #fff;
        }
        .chart-stat .label {
            font-size: 11px;
            color: #888;
            margin-top: 2px;
            text-transform: uppercase;
        }
        .card-full {
            max-width: 900px;
        }
        .report-table-container {
            max-height: 500px;
            overflow-y: auto;
            border-radius: 8px;
            border: 1px solid #21262d;
        }
        .report-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
        }
        .report-table thead { position: sticky; top: 0; z-index: 1; }
        .report-table th {
            background: #0f3460;
            padding: 12px;
            text-align: left;
            font-weight: 600;
            color: #ccc;
            font-size: 11px;
            text-transform: uppercase;
        }
        .report-table td {
            padding: 10px 12px;
            border-bottom: 1px solid #1a1a2e;
        }
        .report-table tr { background: #0d1117; }
        .report-table tr:nth-child(even) { background: #111827; }
        .empty-report {
            text-align: center;
            padding: 40px;
            color: #666;
            font-size: 14px;
        }
    </style>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
</head>
<body>
    <h1>Automacao de Baixa NF</h1>
    <p class="subtitle">Dealer.net - Grupo Indiana</p>

    <!-- TABS PRINCIPAIS -->
    <div class="tabs" style="border-bottom:3px solid #e94560;">
        <button class="tab active" onclick="switchMainTab('mandarim')" id="mainTab-mandarim" style="font-size:16px;">Mandarim</button>
        <button class="tab" onclick="switchMainTab('gaulesa')" id="mainTab-gaulesa" style="font-size:16px;">Gaulesa</button>
    </div>

    <!-- ==================== MANDARIM ==================== -->
    <div class="tab-content active" id="mainContent-mandarim" style="width:100%;">

    <div class="tabs">
        <button class="tab active" onclick="switchTab('painel')" id="subTab-painel">Painel de Controle</button>
        <button class="tab" onclick="switchTab('relatorio')" id="subTab-relatorio">Relatorio e Graficos</button>
        <button class="tab" onclick="switchTab('analise')" id="subTab-analise">Analise (sem baixar)</button>
    </div>

    <!-- ABA 1: PAINEL -->
    <div class="tab-content active" id="tab-painel">

    <!-- PASSO 1: Selecionar Loja -->
    <div class="card">
        <h2><span class="step">1</span> Selecionar Loja e Excel</h2>
        <select id="loja">
            <option value="">-- Selecione a loja --</option>
            <option value="mandarim_iguatemi">Mandarim Iguatemi</option>
            <option value="mandarim_itabuna">Mandarim Itabuna</option>
            <option value="mandarim_lauro">Mandarim Lauro de Freitas</option>
        </select>
        <div class="file-upload">
            <label for="arquivoExcel" class="file-label">
                <span id="arquivoLabel">Clique para selecionar o Excel</span>
            </label>
            <input type="file" id="arquivoExcel" accept=".xlsx,.xls" style="display:none;">
        </div>
        <div id="info-excel" style="font-size:13px; color:#888; margin-bottom:16px;"></div>
        <button class="btn btn-primary" id="btnConfigurar" onclick="configurarDealer()">
            Configurar Dealer
        </button>
    </div>

    <!-- PASSO 2: Instrução -->
    <div class="card" id="card-instrucao" style="display:none;">
        <h2><span class="step">2</span> Preparar o Dealer.net</h2>
        <div class="aviso">
            O navegador foi aberto. Agora faca manualmente:<br><br>
            1. Faca login no Dealer.net<br>
            2. Selecione a loja <strong id="loja-nome-instrucao"></strong><br>
            3. Va em <strong>Financeiro > Contas a Receber > Titulo a Receber</strong><br>
            4. Expanda o <strong>Filtro Avancado</strong><br><br>
            Quando tudo estiver pronto, clique no botao abaixo.
        </div>
        <button class="btn btn-success" id="btnComecar" onclick="comecarBaixas()">
            Comecar Baixas
        </button>
    </div>

    <!-- PASSO 3: Confirmação -->
    <div class="card" id="card-confirma" style="display:none;">
        <h2><span class="step">3</span> Iniciar Baixas</h2>
        <div class="aviso" style="border-left-color: #0cca4a;">
            O Dealer.net foi detectado! Antes de iniciar, confirme:<br><br>
            - A loja correta esta selecionada no Dealer?<br>
            - O Filtro de Selecao esta correto?<br>
            - Tudo pronto para comecar?<br><br>
            <strong>So clique quando estiver TUDO configurado.</strong>
        </div>
        <button class="btn btn-success" id="btnIniciar" onclick="iniciarBaixas()" style="font-size:20px; padding:18px;">
            INICIAR BAIXAS
        </button>
    </div>

    <!-- PASSO 4: Progresso -->
    <div class="card" id="card-progresso" style="display:none;">
        <h2><span class="step">3</span> Progresso das Baixas</h2>
        <div class="nf-atual" id="nf-atual"></div>
        <div class="progress-bar-container">
            <div class="progress-bar" id="progressBar"></div>
        </div>
        <div class="status-bar">
            <div class="status-item total">
                <div class="number" id="stat-total">0</div>
                <div class="label">TOTAL</div>
            </div>
            <div class="status-item ok">
                <div class="number" id="stat-sucesso">0</div>
                <div class="label">SUCESSO</div>
            </div>
            <div class="status-item pago">
                <div class="number" id="stat-pago">0</div>
                <div class="label">JA PAGAS</div>
            </div>
            <div class="status-item erro">
                <div class="number" id="stat-erro">0</div>
                <div class="label">ERROS</div>
            </div>
        </div>
        <!-- Tabela de Monitoramento -->
        <div class="nf-table-container" id="tabelaContainer">
            <table class="nf-table">
                <thead>
                    <tr>
                        <th style="width:50px">#</th>
                        <th style="width:80px">NF</th>
                        <th>Status</th>
                        <th>Detalhe</th>
                    </tr>
                </thead>
                <tbody id="tabelaBody"></tbody>
            </table>
        </div>

        <details style="margin-bottom:16px;">
            <summary style="cursor:pointer; color:#888; font-size:12px;">Log detalhado</summary>
            <div class="log-box" id="logBox" style="margin-top:8px;"></div>
        </details>

        <div style="display:flex; gap:10px;">
            <button class="btn btn-primary" id="btnPausar" onclick="pausarBaixas()" style="flex:1;">
                Pausar Baixas
            </button>
            <button class="btn btn-danger" id="btnParar" onclick="pararBaixas()" style="flex:1;">
                Parar Definitivo
            </button>
        </div>
        <button class="btn btn-success" id="btnExportar" onclick="exportarExcel()" style="margin-top:10px; background:#217346;">
            Exportar Relatorio em Excel
        </button>
    </div>

    </div> <!-- /tab-painel -->

    <!-- ABA 2: RELATORIO -->
    <div class="tab-content" id="tab-relatorio">

        <!-- Grafico de Pizza -->
        <div class="chart-container">
            <h2 style="font-size:18px; color:#e94560; margin-bottom:20px; text-align:center;">Distribuicao dos Resultados</h2>
            <div class="chart-wrapper">
                <canvas id="pieChart"></canvas>
            </div>
            <div class="chart-stats" id="chartStats"></div>
        </div>

        <!-- Tabela detalhada -->
        <div class="card card-full">
            <h2 style="font-size:18px; color:#e94560; margin-bottom:20px;">Detalhamento das Notas</h2>
            <div class="report-table-container" id="reportTableContainer">
                <div class="empty-report">Nenhum dado ainda. Execute uma rodagem primeiro.</div>
            </div>
            <button class="btn btn-success" onclick="exportarExcel()" style="margin-top:16px; background:#217346;">
                Exportar para Excel
            </button>
        </div>

    </div> <!-- /tab-relatorio -->

    <!-- ABA 3: ANALISE -->
    <div class="tab-content" id="tab-analise">

        <!-- PASSO 1: Selecionar Loja (Analise) -->
        <div class="card">
            <h2><span class="step">1</span> Selecionar Loja e Excel</h2>
            <select id="lojaAnalise">
                <option value="">-- Selecione a loja --</option>
                <option value="mandarim_iguatemi">Mandarim Iguatemi</option>
                <option value="mandarim_itabuna">Mandarim Itabuna</option>
                <option value="mandarim_lauro">Mandarim Lauro de Freitas</option>
            </select>
            <div class="file-upload">
                <label for="arquivoExcelAnalise" class="file-label">
                    <span id="arquivoLabelAnalise">Clique para selecionar o Excel</span>
                </label>
                <input type="file" id="arquivoExcelAnalise" accept=".xlsx,.xls" style="display:none;">
            </div>
            <div id="info-excel-analise" style="font-size:13px; color:#888; margin-bottom:16px;"></div>
            <button class="btn btn-primary" id="btnConfigurarAnalise" onclick="configurarAnalise()">
                Configurar Dealer (Analise)
            </button>
        </div>

        <!-- PASSO 2: Iniciar Analise -->
        <div class="card" id="card-comecar-analise" style="display:none;">
            <h2><span class="step">2</span> Abrir Navegador</h2>
            <div class="aviso">
                Clique abaixo para abrir o navegador e iniciar a conexao com o Dealer.<br>
                Depois faca login, selecione a loja e va em Titulo a Receber.
            </div>
            <button class="btn btn-success" id="btnComecarAnalise" onclick="comecarAnalise()">
                Abrir Navegador
            </button>
        </div>

        <!-- PASSO 3: Confirmar Analise -->
        <div class="card" id="card-confirma-analise" style="display:none;">
            <h2><span class="step">3</span> Iniciar Analise</h2>
            <div class="aviso" style="border-left-color: #0cca4a;">
                O Dealer foi detectado! Configure o Filtro de Selecao e a loja correta.<br>
                Esta analise <strong>NAO vai fazer baixa nenhuma</strong> - apenas consultar os valores.
            </div>
            <button class="btn btn-success" id="btnIniciarAnalise" onclick="iniciarAnalise()" style="font-size:20px; padding:18px;">
                INICIAR ANALISE
            </button>
        </div>

        <!-- Progresso e Tabela -->
        <div class="card card-full" id="card-tabela-analise" style="display:none;">
            <h2 style="font-size:18px; color:#e94560; margin-bottom:20px;">Analise de Valores</h2>
            <div class="nf-atual" id="nf-atual-analise"></div>
            <div class="progress-bar-container">
                <div class="progress-bar" id="progressBarAnalise"></div>
            </div>
            <div class="status-bar" id="totais-analise"></div>
            <div class="report-table-container">
                <table class="report-table">
                    <thead>
                        <tr>
                            <th style="width:50px">#</th>
                            <th>NF</th>
                            <th>Valor Total da Nota</th>
                            <th>Saldo</th>
                            <th>Valor Baixa Dealer</th>
                            <th>Valor Baixa Excel</th>
                            <th>Diferenca</th>
                        </tr>
                    </thead>
                    <tbody id="tabelaAnaliseBody"></tbody>
                </table>
            </div>
            <button class="btn btn-success" onclick="exportarAnalise()" style="margin-top:16px; background:#217346;">
                Exportar Analise em Excel
            </button>
        </div>

    </div> <!-- /tab-analise -->

    </div> <!-- /mainContent-mandarim -->

    <!-- ==================== GAULESA ==================== -->
    <div class="tab-content" id="mainContent-gaulesa" style="width:100%;">

    <div class="tabs">
        <button class="tab active" onclick="switchGauleTab('gaulesa_baixa')" id="gauleTab-gaulesa_baixa">Baixa NF</button>
        <button class="tab" onclick="switchGauleTab('gaulesa_cancelar')" id="gauleTab-gaulesa_cancelar">Cancelamento</button>
    </div>

    <div class="tab-content active" id="gauleContent-gaulesa_baixa">

    <!-- PASSO 1: Upload Excel Gaulesa -->
    <div class="card">
        <h2><span class="step">1</span> Upload Excel Gaulesa</h2>
        <div class="file-upload">
            <label for="arquivoGaulesa" class="file-label">
                <span id="arquivoLabelGaulesa">Clique para selecionar o Excel da Gaulesa</span>
            </label>
            <input type="file" id="arquivoGaulesa" accept=".xlsx,.xls" style="display:none;">
        </div>
        <div id="info-gaulesa" style="font-size:13px; color:#888; margin-bottom:16px;"></div>
        <button class="btn btn-primary" id="btnConfigurarGaulesa" onclick="configurarGaulesa()">
            Configurar Gaulesa
        </button>
    </div>

    <!-- PASSO 2: Abrir Navegador -->
    <div class="card" id="card-comecar-gaulesa" style="display:none;">
        <h2><span class="step">2</span> Abrir Navegador</h2>
        <div class="aviso">
            Clique abaixo para abrir o navegador.<br>
            Depois faca login, selecione <strong>GAULESA IGUATEMI</strong> e va em Titulo a Receber.
        </div>
        <button class="btn btn-success" id="btnComecarGaulesa" onclick="comecarGaulesa()">
            Abrir Navegador
        </button>
    </div>

    <!-- PASSO 3: Confirmar -->
    <div class="card" id="card-confirma-gaulesa" style="display:none;">
        <h2><span class="step">3</span> Iniciar Baixas Gaulesa</h2>
        <div class="aviso" style="border-left-color: #0cca4a;">
            Dealer detectado! Busca por <strong>CHASSI</strong> com match por <strong>VALOR</strong>.<br>
            Verifique se a loja esta correta e clique para iniciar.
        </div>
        <button class="btn btn-success" id="btnIniciarGaulesa" onclick="iniciarGaulesa()" style="font-size:20px; padding:18px;">
            INICIAR BAIXAS GAULESA
        </button>
    </div>

    <!-- PASSO 4: Progresso -->
    <div class="card card-full" id="card-progresso-gaulesa" style="display:none;">
        <h2><span class="step">4</span> Progresso Gaulesa</h2>
        <div class="nf-atual" id="nf-atual-gaulesa"></div>
        <div class="progress-bar-container">
            <div class="progress-bar" id="progressBarGaulesa"></div>
        </div>
        <div class="status-bar">
            <div class="status-item total"><div class="number" id="stat-total-g">0</div><div class="label">TOTAL</div></div>
            <div class="status-item ok"><div class="number" id="stat-sucesso-g">0</div><div class="label">SUCESSO</div></div>
            <div class="status-item pago"><div class="number" id="stat-pago-g">0</div><div class="label">JA PAGAS</div></div>
            <div class="status-item erro"><div class="number" id="stat-erro-g">0</div><div class="label">ERROS</div></div>
        </div>
        <div class="nf-table-container" id="tabelaContainerGaulesa">
            <table class="nf-table">
                <thead><tr>
                    <th style="width:50px">#</th>
                    <th>Chassi</th>
                    <th>Valor Excel</th>
                    <th>Status</th>
                    <th>Detalhe</th>
                </tr></thead>
                <tbody id="tabelaBodyGaulesa"></tbody>
            </table>
        </div>
        <div style="display:flex; gap:10px; margin-top:10px;">
            <button class="btn btn-primary" id="btnPausarG" onclick="pausarGaulesa()" style="flex:1;">Pausar</button>
            <button class="btn btn-danger" id="btnPararG" onclick="pararGaulesa()" style="flex:1;">Parar</button>
        </div>
        <button class="btn btn-success" onclick="exportarExcel()" style="margin-top:10px; background:#217346;">
            Exportar Relatorio em Excel
        </button>
    </div>

    </div> <!-- /gauleContent-gaulesa_baixa -->

    <div class="tab-content" id="gauleContent-gaulesa_cancelar">

    <!-- Cancelamento Gaulesa -->
    <div class="card" style="border-color:#dc3545;">
        <h2 style="color:#dc3545;">Cancelar Notas Gaulesa</h2>
        <div class="aviso" style="border-left-color:#dc3545;">
            Use para cancelar baixas feitas incorretamente.<br>
            Busca por Chassi, encontra a nota pelo Valor e cancela com motivo "Erro".
        </div>
        <div class="file-upload">
            <label for="arquivoCancelar" class="file-label">
                <span id="arquivoLabelCancelar">Clique para selecionar o Excel para cancelar</span>
            </label>
            <input type="file" id="arquivoCancelar" accept=".xlsx,.xls" style="display:none;">
        </div>
        <button class="btn btn-danger" id="btnConfigurarCancelar" onclick="configurarCancelamento()" style="margin-top:10px;">
            Configurar Cancelamento
        </button>
        <div id="card-comecar-cancelar" style="display:none; margin-top:16px;">
            <button class="btn btn-danger" id="btnComecarCancelar" onclick="comecarCancelamento()" style="font-size:18px; padding:16px;">
                Abrir Navegador para Cancelamento
            </button>
        </div>
        <div id="card-confirma-cancelar" style="display:none; margin-top:16px;">
            <div class="aviso" style="border-left-color:#dc3545;">
                Dealer detectado! Configure a loja e clique para iniciar o cancelamento.
            </div>
            <button class="btn btn-danger" id="btnIniciarCancelar" onclick="iniciarCancelamento()" style="font-size:20px; padding:18px;">
                INICIAR CANCELAMENTO
            </button>
        </div>
        <div id="card-progresso-cancelar" style="display:none; margin-top:16px;">
            <div class="nf-atual" id="nf-atual-cancelar"></div>
            <div class="progress-bar-container">
                <div class="progress-bar" id="progressBarCancelar" style="background:linear-gradient(90deg,#dc3545,#ff8500);"></div>
            </div>
            <div class="nf-table-container" id="tabelaContainerCancelar">
                <table class="nf-table">
                    <thead><tr>
                        <th>#</th><th>Chassi</th><th>Valor</th><th>Status</th><th>Detalhe</th>
                    </tr></thead>
                    <tbody id="tabelaBodyCancelar"></tbody>
                </table>
            </div>
            <div style="display:flex; gap:10px; margin-top:10px;">
                <button class="btn btn-primary" id="btnPausarCancel" onclick="pausarBaixas()" style="flex:1;">Pausar</button>
                <button class="btn btn-danger" id="btnPararCancel" onclick="pararBaixas()" style="flex:1;">Parar</button>
            </div>
        </div>
    </div>

    </div> <!-- /gauleContent-gaulesa_cancelar -->

    </div> <!-- /mainContent-gaulesa -->

    <script>
        function switchGauleTab(tabName) {
            document.querySelectorAll('[id^="gauleTab-"]').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('[id^="gauleContent-"]').forEach(t => t.classList.remove('active'));
            document.getElementById('gauleTab-' + tabName).classList.add('active');
            document.getElementById('gauleContent-' + tabName).classList.add('active');
        }

        function switchMainTab(tabName) {
            document.querySelectorAll('[id^="mainTab-"]').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('[id^="mainContent-"]').forEach(t => t.classList.remove('active'));
            document.getElementById('mainTab-' + tabName).classList.add('active');
            document.getElementById('mainContent-' + tabName).classList.add('active');
        }
        let pieChart = null;

        function switchTab(tabName) {
            document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
            event.target.classList.add('active');
            document.getElementById('tab-' + tabName).classList.add('active');
            if (tabName === 'relatorio') {
                atualizarRelatorio();
            }
        }

        async function atualizarRelatorio() {
            try {
                const resp = await fetch('/api/status');
                const data = await resp.json();
                const tabela = data.tabela_nfs || [];

                // Contadores
                const counts = {sucesso: 0, pago: 0, erro: 0, nao_encontrada: 0};
                tabela.forEach(item => {
                    if (counts[item.status] !== undefined) counts[item.status]++;
                });

                // Atualiza grafico (cria uma vez, depois só atualiza os dados)
                const novoData = [counts.sucesso, counts.pago, counts.erro, counts.nao_encontrada];

                if (!pieChart) {
                    const ctx = document.getElementById('pieChart').getContext('2d');
                    pieChart = new Chart(ctx, {
                        type: 'pie',
                        data: {
                            labels: ['Baixadas', 'Ja Pagas', 'Erros', 'Nao Encontradas'],
                            datasets: [{
                                data: novoData,
                                backgroundColor: ['#0cca4a', '#ffc107', '#dc3545', '#6c757d'],
                                borderColor: '#16213e',
                                borderWidth: 3
                            }]
                        },
                        options: {
                            responsive: true,
                            maintainAspectRatio: true,
                            animation: { duration: 0 },
                            plugins: {
                                legend: {
                                    position: 'bottom',
                                    labels: { color: '#ccc', font: {size: 13}, padding: 15 }
                                },
                                tooltip: {
                                    callbacks: {
                                        label: function(ctx) {
                                            const total = ctx.dataset.data.reduce((a,b)=>a+b,0);
                                            const pct = total > 0 ? ((ctx.parsed/total)*100).toFixed(1) : 0;
                                            return ctx.label + ': ' + ctx.parsed + ' (' + pct + '%)';
                                        }
                                    }
                                }
                            }
                        }
                    });
                } else {
                    // Só atualiza os dados sem recriar
                    const atual = pieChart.data.datasets[0].data;
                    const mudou = atual.some((v, i) => v !== novoData[i]);
                    if (mudou) {
                        pieChart.data.datasets[0].data = novoData;
                        pieChart.update('none');
                    }
                }

                // Stats cards
                const statsHtml = [
                    {label: 'Baixadas', count: counts.sucesso, color: '#0cca4a'},
                    {label: 'Ja Pagas', count: counts.pago, color: '#ffc107'},
                    {label: 'Erros', count: counts.erro, color: '#dc3545'},
                    {label: 'Nao Encontradas', count: counts.nao_encontrada, color: '#6c757d'}
                ].map(s =>
                    '<div class="chart-stat">' +
                    '<div><span class="dot" style="background:' + s.color + '"></span>' +
                    '<span class="number">' + s.count + '</span></div>' +
                    '<div class="label">' + s.label + '</div>' +
                    '</div>'
                ).join('');
                document.getElementById('chartStats').innerHTML = statsHtml;

                // Tabela
                const container = document.getElementById('reportTableContainer');
                if (tabela.length === 0) {
                    container.innerHTML = '<div class="empty-report">Nenhum dado ainda. Execute uma rodagem primeiro.</div>';
                    return;
                }

                const STATUS_LABELS = {
                    'sucesso': 'Baixada',
                    'pago': 'Ja Paga',
                    'erro': 'Erro',
                    'nao_encontrada': 'Nao Encontrada',
                    'processando': 'Processando'
                };

                const rows = tabela.map((item, idx) => {
                    const valor = Number(item.valor || 0).toLocaleString('pt-BR', {
                        style: 'currency', currency: 'BRL'
                    });
                    const valorTotal = item.valor_total_nota
                        ? 'R$ ' + item.valor_total_nota
                        : '-';
                    const badge = '<span class="badge badge-' + item.status + '">' + (STATUS_LABELS[item.status] || item.status) + '</span>';
                    return '<tr>' +
                        '<td style="color:#555">' + (idx + 1) + '</td>' +
                        '<td>' + (item.cnpj || '-') + '</td>' +
                        '<td style="font-weight:600; color:#ccc">' + (item.nf_original || item.nf) + '</td>' +
                        '<td style="color:#0cca4a; font-weight:600">' + valor + '</td>' +
                        '<td style="color:#58a6ff; font-weight:600">' + valorTotal + '</td>' +
                        '<td>' + badge + '</td>' +
                        '</tr>';
                }).join('');

                container.innerHTML =
                    '<table class="report-table"><thead><tr>' +
                    '<th style="width:50px">#</th>' +
                    '<th>CNPJ</th>' +
                    '<th>NF</th>' +
                    '<th>Total da Baixa</th>' +
                    '<th>Valor Total da Nota</th>' +
                    '<th>Status</th>' +
                    '</tr></thead><tbody>' + rows + '</tbody></table>';
            } catch(e) {
                console.error(e);
            }
        }

        // Auto-atualiza o relatório a cada 3s se estiver na aba
        setInterval(() => {
            if (document.getElementById('tab-relatorio').classList.contains('active')) {
                atualizarRelatorio();
            }
            if (document.getElementById('tab-analise').classList.contains('active')) {
                atualizarTabelaAnalise();
            }
            if (document.getElementById('mainContent-gaulesa').classList.contains('active')) {
                atualizarProgressoGaulesa();
                atualizarProgressoCancelamento();
            }
        }, 1500);

        // ============================================================
        // GAULESA
        // ============================================================
        document.getElementById('arquivoGaulesa').addEventListener('change', function() {
            const label = document.getElementById('arquivoLabelGaulesa');
            if (this.files && this.files[0]) {
                label.textContent = '✓ ' + this.files[0].name;
                this.previousElementSibling.classList.add('has-file');
            }
        });

        async function configurarGaulesa() {
            const arquivo = document.getElementById('arquivoGaulesa').files[0];
            if (!arquivo) { alert('Selecione o arquivo Excel da Gaulesa!'); return; }
            const btn = document.getElementById('btnConfigurarGaulesa');
            btn.disabled = true;
            btn.textContent = 'Enviando...';
            try {
                const formData = new FormData();
                formData.append('arquivo', arquivo);
                const resp = await fetch('/api/configurar_gaulesa', {method: 'POST', body: formData});
                const data = await resp.json();
                if (data.ok) {
                    document.getElementById('card-comecar-gaulesa').style.display = 'block';
                    btn.textContent = 'Configurado (' + data.total + ' chassis)';
                } else {
                    alert('Erro: ' + data.erro);
                    btn.disabled = false;
                    btn.textContent = 'Configurar Gaulesa';
                }
            } catch(e) {
                alert('Erro: ' + e.message);
                btn.disabled = false;
                btn.textContent = 'Configurar Gaulesa';
            }
        }

        async function comecarGaulesa() {
            const btn = document.getElementById('btnComecarGaulesa');
            btn.disabled = true;
            btn.textContent = 'Abrindo navegador...';
            try {
                const resp = await fetch('/api/comecar_gaulesa', {method: 'POST'});
                const data = await resp.json();
                if (data.ok) {
                    document.getElementById('stat-total-g').textContent = data.total;
                    btn.textContent = 'Aguardando Dealer...';
                    const check = setInterval(async () => {
                        try {
                            const r = await fetch('/api/status');
                            const d = await r.json();
                            if (d.dealer_pronto) {
                                clearInterval(check);
                                document.getElementById('card-confirma-gaulesa').style.display = 'block';
                                btn.textContent = 'Dealer Detectado!';
                            }
                        } catch(e) {}
                    }, 2000);
                } else {
                    alert('Erro: ' + data.erro);
                    btn.disabled = false;
                    btn.textContent = 'Abrir Navegador';
                }
            } catch(e) {
                alert('Erro: ' + e.message);
                btn.disabled = false;
                btn.textContent = 'Abrir Navegador';
            }
        }

        async function iniciarGaulesa() {
            const btn = document.getElementById('btnIniciarGaulesa');
            btn.disabled = true;
            btn.textContent = 'Iniciando...';
            await fetch('/api/iniciar', {method: 'POST'});
            document.getElementById('card-progresso-gaulesa').style.display = 'block';
        }

        let pausadoG = false;
        async function pausarGaulesa() {
            const btn = document.getElementById('btnPausarG');
            if (!pausadoG) {
                await fetch('/api/pausar', {method: 'POST'});
                pausadoG = true;
                btn.textContent = 'Recomecar';
                btn.className = 'btn btn-success';
                btn.style.flex = '1';
            } else {
                await fetch('/api/recomecar', {method: 'POST'});
                pausadoG = false;
                btn.textContent = 'Pausar';
                btn.className = 'btn btn-primary';
                btn.style.flex = '1';
            }
        }
        async function pararGaulesa() {
            await fetch('/api/parar', {method: 'POST'});
        }

        // ============================================================
        // CANCELAMENTO GAULESA
        // ============================================================
        document.getElementById('arquivoCancelar').addEventListener('change', function() {
            const label = document.getElementById('arquivoLabelCancelar');
            if (this.files && this.files[0]) {
                label.textContent = '✓ ' + this.files[0].name;
                this.previousElementSibling.classList.add('has-file');
            }
        });

        async function atualizarProgressoCancelamento() {
            const container = document.getElementById('card-progresso-cancelar');
            if (container.style.display === 'none') return;
            try {
                const resp = await fetch('/api/status');
                const data = await resp.json();
                const p = data.progresso;
                const processadas = p.sucesso + p.nao_encontrada + p.erro;
                const pct = p.total > 0 ? (processadas / p.total * 100) : 0;
                document.getElementById('progressBarCancelar').style.width = pct + '%';
                if (data.nf_atual && data.rodando) {
                    document.getElementById('nf-atual-cancelar').textContent = 'Cancelando: ' + data.nf_atual + ' (' + processadas + '/' + p.total + ')';
                } else if (!data.rodando && processadas > 0) {
                    document.getElementById('nf-atual-cancelar').textContent = 'CONCLUIDO! ' + processadas + '/' + p.total;
                }
                const tabela = data.tabela_nfs || [];
                document.getElementById('tabelaBodyCancelar').innerHTML = tabela.map((item, idx) => {
                    const valor = Number(item.valor || 0).toLocaleString('pt-BR', {style:'currency', currency:'BRL'});
                    const badge = '<span class="badge badge-' + item.status + '">' + (badgeLabels[item.status] || item.status) + '</span>';
                    return '<tr><td>' + (idx+1) + '</td><td style="font-size:11px;color:#ccc">' + item.nf + '</td><td style="color:#dc3545">' + valor + '</td><td>' + badge + '</td><td style="color:#888;font-size:12px">' + (item.mensagem||'') + '</td></tr>';
                }).join('');
            } catch(e) {}
        }

        async function configurarCancelamento() {
            const arquivo = document.getElementById('arquivoCancelar').files[0];
            if (!arquivo) { alert('Selecione o Excel!'); return; }
            const btn = document.getElementById('btnConfigurarCancelar');
            btn.disabled = true;
            btn.textContent = 'Enviando...';
            try {
                const formData = new FormData();
                formData.append('arquivo', arquivo);
                const resp = await fetch('/api/configurar_cancelamento', {method: 'POST', body: formData});
                const data = await resp.json();
                if (data.ok) {
                    document.getElementById('card-comecar-cancelar').style.display = 'block';
                    btn.textContent = 'Configurado (' + data.total + ' para cancelar)';
                } else {
                    alert('Erro: ' + data.erro);
                    btn.disabled = false;
                    btn.textContent = 'Configurar Cancelamento';
                }
            } catch(e) {
                alert('Erro: ' + e.message);
                btn.disabled = false;
                btn.textContent = 'Configurar Cancelamento';
            }
        }

        async function comecarCancelamento() {
            const btn = document.getElementById('btnComecarCancelar');
            btn.disabled = true;
            btn.textContent = 'Abrindo navegador...';
            try {
                const resp = await fetch('/api/comecar_cancelamento', {method: 'POST'});
                const data = await resp.json();
                if (data.ok) {
                    btn.textContent = 'Aguardando Dealer...';
                    const check = setInterval(async () => {
                        try {
                            const r = await fetch('/api/status');
                            const d = await r.json();
                            if (d.dealer_pronto) {
                                clearInterval(check);
                                document.getElementById('card-confirma-cancelar').style.display = 'block';
                                btn.textContent = 'Dealer Detectado!';
                            }
                        } catch(e) {}
                    }, 2000);
                } else {
                    alert('Erro: ' + data.erro);
                    btn.disabled = false;
                    btn.textContent = 'Abrir Navegador para Cancelamento';
                }
            } catch(e) {
                alert('Erro: ' + e.message);
                btn.disabled = false;
                btn.textContent = 'Abrir Navegador para Cancelamento';
            }
        }

        async function iniciarCancelamento() {
            const btn = document.getElementById('btnIniciarCancelar');
            btn.disabled = true;
            btn.textContent = 'Cancelando...';
            await fetch('/api/iniciar', {method: 'POST'});
            document.getElementById('card-progresso-cancelar').style.display = 'block';
        }

        async function atualizarProgressoGaulesa() {
            try {
                const resp = await fetch('/api/status');
                const data = await resp.json();
                const p = data.progresso;
                const processadas = p.sucesso + p.pago + p.nao_encontrada + p.erro;

                document.getElementById('stat-total-g').textContent = p.total;
                document.getElementById('stat-sucesso-g').textContent = p.sucesso;
                document.getElementById('stat-pago-g').textContent = p.pago;
                document.getElementById('stat-erro-g').textContent = p.erro + p.nao_encontrada;

                const pct = p.total > 0 ? (processadas / p.total * 100) : 0;
                document.getElementById('progressBarGaulesa').style.width = pct + '%';

                if (data.nf_atual && data.rodando) {
                    document.getElementById('nf-atual-gaulesa').textContent =
                        'Chassi: ' + data.nf_atual + ' (' + processadas + '/' + p.total + ')';
                } else if (!data.rodando && processadas > 0) {
                    document.getElementById('nf-atual-gaulesa').textContent =
                        'CONCLUIDO! ' + processadas + '/' + p.total + ' chassis processados';
                }

                // Tabela
                const tabela = data.tabela_nfs || [];
                const tbody = document.getElementById('tabelaBodyGaulesa');
                tbody.innerHTML = tabela.map((item, idx) => {
                    const valor = Number(item.valor || 0).toLocaleString('pt-BR', {style:'currency', currency:'BRL'});
                    const badge = '<span class="badge badge-' + item.status + '">' + (badgeLabels[item.status] || item.status) + '</span>';
                    return '<tr>' +
                        '<td style="color:#555">' + (idx+1) + '</td>' +
                        '<td style="font-weight:600;color:#ccc;font-size:11px">' + item.nf + '</td>' +
                        '<td style="color:#0cca4a">' + valor + '</td>' +
                        '<td>' + badge + '</td>' +
                        '<td style="color:#888;font-size:12px">' + (item.mensagem || '') + '</td>' +
                        '</tr>';
                }).join('');
            } catch(e) {}
        }

        // ============================================================
        // ABA ANALISE
        // ============================================================
        document.getElementById('lojaAnalise').addEventListener('change', function() {
            const info = document.getElementById('info-excel-analise');
            info.textContent = this.value ? 'Agora selecione o arquivo Excel abaixo' : '';
        });

        document.getElementById('arquivoExcelAnalise').addEventListener('change', function() {
            const label = document.getElementById('arquivoLabelAnalise');
            if (this.files && this.files[0]) {
                label.textContent = '✓ ' + this.files[0].name;
                this.previousElementSibling.classList.add('has-file');
            } else {
                label.textContent = 'Clique para selecionar o Excel';
                this.previousElementSibling.classList.remove('has-file');
            }
        });

        async function configurarAnalise() {
            const loja = document.getElementById('lojaAnalise').value;
            const arquivo = document.getElementById('arquivoExcelAnalise').files[0];
            if (!loja) { alert('Selecione uma loja primeiro!'); return; }
            if (!arquivo) { alert('Selecione o arquivo Excel!'); return; }
            const btn = document.getElementById('btnConfigurarAnalise');
            btn.disabled = true;
            btn.textContent = 'Enviando arquivo...';
            try {
                const formData = new FormData();
                formData.append('loja', loja);
                formData.append('arquivo', arquivo);
                const resp = await fetch('/api/configurar', {
                    method: 'POST',
                    body: formData
                });
                const data = await resp.json();
                if (data.ok) {
                    document.getElementById('card-comecar-analise').style.display = 'block';
                    btn.textContent = 'Configurado (' + data.total + ' NFs)';
                } else {
                    alert('Erro: ' + data.erro);
                    btn.disabled = false;
                    btn.textContent = 'Configurar Dealer (Analise)';
                }
            } catch(e) {
                alert('Erro: ' + e.message);
                btn.disabled = false;
                btn.textContent = 'Configurar Dealer (Analise)';
            }
        }

        async function comecarAnalise() {
            const btn = document.getElementById('btnComecarAnalise');
            btn.disabled = true;
            btn.textContent = 'Abrindo navegador...';
            try {
                const resp = await fetch('/api/comecar_analise', {method: 'POST'});
                const data = await resp.json();
                if (data.ok) {
                    btn.textContent = 'Aguardando Dealer...';
                    aguardarDealerProntoAnalise();
                } else {
                    alert('Erro: ' + data.erro);
                    btn.disabled = false;
                    btn.textContent = 'Abrir Navegador';
                }
            } catch(e) {
                alert('Erro: ' + e.message);
                btn.disabled = false;
                btn.textContent = 'Abrir Navegador';
            }
        }

        function aguardarDealerProntoAnalise() {
            const check = setInterval(async () => {
                try {
                    const resp = await fetch('/api/status');
                    const data = await resp.json();
                    if (data.dealer_pronto) {
                        clearInterval(check);
                        document.getElementById('card-confirma-analise').style.display = 'block';
                        document.getElementById('btnComecarAnalise').textContent = 'Dealer Detectado!';
                    }
                } catch(e) {}
            }, 2000);
        }

        async function iniciarAnalise() {
            const btn = document.getElementById('btnIniciarAnalise');
            btn.disabled = true;
            btn.textContent = 'Iniciando...';
            await fetch('/api/iniciar', {method: 'POST'});
            document.getElementById('card-tabela-analise').style.display = 'block';
        }

        function exportarAnalise() {
            window.location.href = '/api/exportar_analise';
        }

        function formatarMoeda(valor) {
            return Number(valor || 0).toLocaleString('pt-BR', {
                style: 'currency', currency: 'BRL'
            });
        }

        function parseValorBR(str) {
            if (!str) return 0;
            try {
                return parseFloat(String(str).replace(/\./g, '').replace(',', '.')) || 0;
            } catch(e) { return 0; }
        }

        async function atualizarTabelaAnalise() {
            try {
                const resp = await fetch('/api/status');
                const data = await resp.json();
                const tabela = data.tabela_analise || [];
                const total = data.progresso.total || 0;
                const processadas = data.progresso.processadas || 0;

                // NF atual
                if (data.nf_atual && data.rodando) {
                    document.getElementById('nf-atual-analise').textContent =
                        'Analisando NF: ' + data.nf_atual + ' (' + processadas + '/' + total + ')';
                } else if (tabela.length > 0 && !data.rodando) {
                    document.getElementById('nf-atual-analise').textContent =
                        'CONCLUIDO! ' + tabela.length + ' NFs analisadas';
                }

                // Barra de progresso
                const pct = total > 0 ? (processadas / total * 100) : 0;
                document.getElementById('progressBarAnalise').style.width = pct + '%';

                // Totalizadores
                let totalDealer = 0, totalExcel = 0, totalNotas = 0;
                tabela.forEach(item => {
                    totalDealer += Number(item.valor_baixa_dealer || 0);
                    totalExcel += Number(item.valor_baixa_excel || 0);
                    totalNotas += parseValorBR(item.valor_total);
                });
                const diferenca = totalDealer - totalExcel;
                const corDif = Math.abs(diferenca) < 0.01 ? '#0cca4a' : '#dc3545';

                document.getElementById('totais-analise').innerHTML =
                    '<div class="status-item"><div class="number" style="color:#17a2b8">' + formatarMoeda(totalNotas) + '</div><div class="label">Total Notas</div></div>' +
                    '<div class="status-item"><div class="number" style="color:#0cca4a">' + formatarMoeda(totalDealer) + '</div><div class="label">Baixa Dealer</div></div>' +
                    '<div class="status-item"><div class="number" style="color:#ffc107">' + formatarMoeda(totalExcel) + '</div><div class="label">Baixa Excel</div></div>' +
                    '<div class="status-item"><div class="number" style="color:' + corDif + '">' + formatarMoeda(diferenca) + '</div><div class="label">Diferenca</div></div>';

                // Tabela
                const tbody = document.getElementById('tabelaAnaliseBody');
                tbody.innerHTML = tabela.map((item, idx) => {
                    const baixaDealer = Number(item.valor_baixa_dealer || 0);
                    const baixaExcel = Number(item.valor_baixa_excel || 0);
                    const dif = baixaDealer - baixaExcel;
                    const corRow = Math.abs(dif) < 0.01 ? '' : 'style="background:#3d1a1a"';
                    return '<tr ' + corRow + '>' +
                        '<td style="color:#555">' + (idx + 1) + '</td>' +
                        '<td style="font-weight:600; color:#ccc">' + (item.nf_original || item.nf) + '</td>' +
                        '<td style="color:#17a2b8">R$ ' + (item.valor_total || '-') + '</td>' +
                        '<td style="color:#ffc107">R$ ' + (item.saldo || '-') + '</td>' +
                        '<td style="color:#0cca4a; font-weight:600">' + formatarMoeda(baixaDealer) + '</td>' +
                        '<td style="color:#ffc107; font-weight:600">' + formatarMoeda(baixaExcel) + '</td>' +
                        '<td style="color:' + (Math.abs(dif)<0.01?'#0cca4a':'#dc3545') + '; font-weight:600">' + formatarMoeda(dif) + '</td>' +
                        '</tr>';
                }).join('');
            } catch(e) { console.error(e); }
        }

        let polling = null;

        document.getElementById('loja').addEventListener('change', function() {
            const info = document.getElementById('info-excel');
            info.textContent = this.value ? 'Agora selecione o arquivo Excel abaixo' : '';
        });

        document.getElementById('arquivoExcel').addEventListener('change', function() {
            const label = document.getElementById('arquivoLabel');
            if (this.files && this.files[0]) {
                label.textContent = '✓ ' + this.files[0].name;
                document.querySelector('#arquivoExcel').previousElementSibling.classList.add('has-file');
            } else {
                label.textContent = 'Clique para selecionar o Excel';
                document.querySelector('#arquivoExcel').previousElementSibling.classList.remove('has-file');
            }
        });

        async function configurarDealer() {
            const loja = document.getElementById('loja').value;
            const arquivo = document.getElementById('arquivoExcel').files[0];
            if (!loja) { alert('Selecione uma loja primeiro!'); return; }
            if (!arquivo) { alert('Selecione o arquivo Excel!'); return; }

            const btn = document.getElementById('btnConfigurar');
            btn.disabled = true;
            btn.textContent = 'Enviando arquivo...';

            try {
                const formData = new FormData();
                formData.append('loja', loja);
                formData.append('arquivo', arquivo);
                const resp = await fetch('/api/configurar', {
                    method: 'POST',
                    body: formData
                });
                const data = await resp.json();
                if (data.ok) {
                    document.getElementById('card-instrucao').style.display = 'block';
                    const nomes = {
                        'mandarim_iguatemi': 'Mandarim Iguatemi',
                        'mandarim_itabuna': 'Mandarim Itabuna',
                        'mandarim_lauro': 'Mandarim Lauro de Freitas'
                    };
                    document.getElementById('loja-nome-instrucao').textContent = nomes[loja];
                    btn.textContent = 'Navegador Aberto';
                } else {
                    alert('Erro: ' + data.erro);
                    btn.disabled = false;
                    btn.textContent = 'Configurar Dealer';
                }
            } catch(e) {
                alert('Erro de conexao: ' + e.message);
                btn.disabled = false;
                btn.textContent = 'Configurar Dealer';
            }
        }

        async function comecarBaixas() {
            const btn = document.getElementById('btnComecar');
            btn.disabled = true;
            btn.textContent = 'Abrindo navegador...';

            try {
                const resp = await fetch('/api/comecar', {method: 'POST'});
                const data = await resp.json();
                if (data.ok) {
                    document.getElementById('stat-total').textContent = data.total;
                    btn.textContent = 'Aguardando Dealer...';
                    // Fica checando até o Dealer estar pronto
                    aguardarDealerPronto();
                } else {
                    alert('Erro: ' + data.erro);
                    btn.disabled = false;
                    btn.textContent = 'Comecar Baixas';
                }
            } catch(e) {
                alert('Erro: ' + e.message);
                btn.disabled = false;
                btn.textContent = 'Comecar Baixas';
            }
        }

        function aguardarDealerPronto() {
            const check = setInterval(async () => {
                try {
                    const resp = await fetch('/api/status');
                    const data = await resp.json();
                    if (data.dealer_pronto) {
                        clearInterval(check);
                        document.getElementById('card-confirma').style.display = 'block';
                        document.getElementById('btnComecar').textContent = 'Dealer Detectado!';
                    }
                } catch(e) {}
            }, 2000);
        }

        async function iniciarBaixas() {
            const btn = document.getElementById('btnIniciar');
            btn.disabled = true;
            btn.textContent = 'Iniciando...';

            await fetch('/api/iniciar', {method: 'POST'});
            document.getElementById('card-progresso').style.display = 'block';
            iniciarPolling();
        }

        let pausado = false;

        async function pausarBaixas() {
            const btn = document.getElementById('btnPausar');
            if (!pausado) {
                await fetch('/api/pausar', {method: 'POST'});
                pausado = true;
                btn.textContent = 'Recomecar Baixas';
                btn.className = 'btn btn-success';
                btn.style.flex = '1';
            } else {
                await fetch('/api/recomecar', {method: 'POST'});
                pausado = false;
                btn.textContent = 'Pausar Baixas';
                btn.className = 'btn btn-primary';
                btn.style.flex = '1';
            }
        }

        function exportarExcel() {
            window.location.href = '/api/exportar';
        }

        async function pararBaixas() {
            try {
                await fetch('/api/parar', {method: 'POST'});
                if (polling) clearInterval(polling);
            } catch(e) {}
        }

        const badgeLabels = {
            'sucesso': 'Baixada',
            'pago': 'Ja Paga',
            'erro': 'Erro',
            'nao_encontrada': 'Nao Encontrada',
            'processando': 'Baixando...',
            'aguardando': 'Aguardando',
            'baixada_anteriormente': 'Baixada Anterior'
        };

        let lastTableLen = 0;

        function iniciarPolling() {
            polling = setInterval(async () => {
                try {
                    const resp = await fetch('/api/status');
                    const data = await resp.json();
                    const p = data.progresso;
                    const processadas = p.sucesso + p.pago + p.nao_encontrada + p.erro;

                    document.getElementById('stat-total').textContent = p.total;
                    document.getElementById('stat-sucesso').textContent = p.sucesso;
                    document.getElementById('stat-pago').textContent = p.pago;
                    document.getElementById('stat-erro').textContent = p.erro + p.nao_encontrada;

                    const pct = p.total > 0 ? (processadas / p.total * 100) : 0;
                    document.getElementById('progressBar').style.width = pct + '%';

                    if (data.nf_atual) {
                        document.getElementById('nf-atual').textContent =
                            'Processando NF: ' + data.nf_atual + ' (' + processadas + '/' + p.total + ')';
                    }

                    // Atualizar tabela de NFs
                    const tbody = document.getElementById('tabelaBody');
                    const tabela = data.tabela_nfs || [];

                    if (tabela.length !== lastTableLen) {
                        // Rebuild table only when new entries arrive
                        tbody.innerHTML = tabela.map((item, idx) => {
                            const isProcessando = item.status === 'processando';
                            const rowClass = isProcessando ? ' class="processando"' : '';
                            const badge = '<span class="badge badge-' + item.status + '">' + (badgeLabels[item.status] || item.status) + '</span>';
                            return '<tr' + rowClass + '>' +
                                '<td style="color:#555">' + (idx + 1) + '</td>' +
                                '<td style="font-weight:600; color:#ccc">' + item.nf + '</td>' +
                                '<td>' + badge + '</td>' +
                                '<td style="color:#888; font-size:12px">' + (item.mensagem || '') + '</td>' +
                                '</tr>';
                        }).join('');
                        lastTableLen = tabela.length;

                        // Auto-scroll para a última linha
                        const container = document.getElementById('tabelaContainer');
                        container.scrollTop = container.scrollHeight;
                    } else if (tabela.length > 0) {
                        // Update last row status (for processando -> result transition)
                        const lastItem = tabela[tabela.length - 1];
                        const lastRow = tbody.lastElementChild;
                        if (lastRow) {
                            const isProcessando = lastItem.status === 'processando';
                            lastRow.className = isProcessando ? 'processando' : '';
                            const badge = '<span class="badge badge-' + lastItem.status + '">' + (badgeLabels[lastItem.status] || lastItem.status) + '</span>';
                            lastRow.cells[2].innerHTML = badge;
                            lastRow.cells[3].innerHTML = '<span style="color:#888; font-size:12px">' + (lastItem.mensagem || '') + '</span>';
                        }
                    }

                    // Atualizar log detalhado
                    const logBox = document.getElementById('logBox');
                    logBox.innerHTML = data.log_mensagens.slice(-50).map(m => {
                        let cls = 'info';
                        if (m.includes('CONFIRMADA') || m.includes('sucesso')) cls = 'success';
                        else if (m.includes('ERRO')) cls = 'error';
                        else if (m.includes('PAGA') || m.includes('pulando') || m.includes('Pago')) cls = 'warning';
                        return '<div class="log-line ' + cls + '">' + m + '</div>';
                    }).join('');
                    logBox.scrollTop = logBox.scrollHeight;

                    if (!data.rodando && processadas > 0) {
                        clearInterval(polling);
                        document.getElementById('nf-atual').textContent = 'CONCLUIDO! ' + processadas + '/' + p.total + ' NFs processadas';
                        document.getElementById('btnParar').textContent = 'Finalizado';
                        document.getElementById('btnParar').disabled = true;
                    }
                } catch(e) {}
            }, 1000);
        }
    </script>
</body>
</html>
"""


@app.route("/")
def index():
    return render_template_string(HTML_PAGE)


@app.route("/api/configurar", methods=["POST"])
def api_configurar():
    global automacao, estado

    # Aceita tanto JSON quanto multipart (com upload de arquivo)
    if request.content_type and "multipart" in request.content_type:
        loja_key = request.form.get("loja")
        arquivo_upload = request.files.get("arquivo")
    else:
        data = request.json or {}
        loja_key = data.get("loja")
        arquivo_upload = None

    if loja_key not in LOJAS:
        return jsonify({"ok": False, "erro": "Loja inválida"})

    # Se veio arquivo pelo upload, salva na pasta uploads
    if arquivo_upload and arquivo_upload.filename:
        if not arquivo_upload.filename.lower().endswith((".xlsx", ".xls")):
            return jsonify({"ok": False, "erro": "Arquivo deve ser .xlsx ou .xls"})
        uploads_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
        os.makedirs(uploads_dir, exist_ok=True)
        # Nome único por loja
        caminho = os.path.join(uploads_dir, f"{loja_key}.xlsx")
        arquivo_upload.save(caminho)
    else:
        # Fallback: usa o arquivo padrão da pasta (se existir)
        loja = LOJAS[loja_key]
        arquivo = loja["arquivo"]
        caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), arquivo)

    if not os.path.exists(caminho):
        return jsonify({"ok": False, "erro": "Nenhum arquivo Excel disponivel. Faca upload do Excel."})

    estado["loja_selecionada"] = loja_key
    estado["log_mensagens"] = []
    estado["rodando"] = False
    estado["dealer_pronto"] = False
    estado["inicio_confirmado"] = False
    estado["browser_aberto"] = False
    estado["tabela_nfs"] = []
    estado["tabela_analise"] = []
    estado["nf_atual"] = ""

    # Cria o objeto da automação
    print(f"[SERVIDOR] Criando automacao com loja_key={loja_key}, caminho={caminho}")
    automacao = AutomacaoBaixa(caminho, estado, loja_key=loja_key)
    total = automacao.carregar_notas()
    print(f"[SERVIDOR] Notas carregadas: {total}")
    estado["browser_aberto"] = True
    return jsonify({"ok": True, "total": total})


@app.route("/api/comecar", methods=["POST"])
def api_comecar():
    global automacao, estado
    if automacao is None:
        return jsonify({"ok": False, "erro": "Configure o Dealer primeiro!"})

    if estado["rodando"]:
        return jsonify({"ok": False, "erro": "Já está rodando!"})

    total = len(automacao.notas)
    estado["progresso"] = {"total": total, "processadas": 0, "sucesso": 0, "pago": 0, "nao_encontrada": 0, "baixada_anteriormente": 0, "erro": 0}
    estado["tabela_nfs"] = []
    estado["log_mensagens"] = []
    estado["rodando"] = True
    estado["dealer_pronto"] = False
    estado["inicio_confirmado"] = False

    # Tudo (browser + automação) roda numa única thread
    thread = threading.Thread(target=automacao.executar_tudo, daemon=True)
    thread.start()

    return jsonify({"ok": True, "total": total})


@app.route("/api/status")
def api_status():
    return jsonify({
        "rodando": estado["rodando"],
        "progresso": estado["progresso"],
        "nf_atual": estado["nf_atual"],
        "log_mensagens": estado["log_mensagens"][-100:],
        "tabela_nfs": estado["tabela_nfs"],
        "tabela_analise": estado["tabela_analise"],
        "dealer_pronto": estado["dealer_pronto"],
    })


@app.route("/api/comecar_analise", methods=["POST"])
def api_comecar_analise():
    global automacao, estado
    if automacao is None:
        return jsonify({"ok": False, "erro": "Configure o Dealer primeiro!"})

    if estado["rodando"]:
        return jsonify({"ok": False, "erro": "Ja esta rodando!"})

    total = len(automacao.notas)
    estado["progresso"] = {"total": total, "processadas": 0, "sucesso": 0, "pago": 0, "nao_encontrada": 0, "baixada_anteriormente": 0, "erro": 0}
    estado["tabela_analise"] = []
    estado["log_mensagens"] = []
    estado["rodando"] = True
    estado["dealer_pronto"] = False
    estado["inicio_confirmado"] = False

    thread = threading.Thread(target=automacao.executar_analise, daemon=True)
    thread.start()

    return jsonify({"ok": True, "total": total})


@app.route("/api/exportar_analise")
def api_exportar_analise():
    """Exporta a tabela de análise em Excel."""
    tabela = estado.get("tabela_analise", [])
    if not tabela:
        return jsonify({"ok": False, "erro": "Nenhum dado para exportar"}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analise"

    headers = ["CNPJ", "NF", "Valor Total da Nota", "Saldo", "Valor da Baixa Dealer", "Valor da Baixa Excel", "Diferenca"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def to_float(s):
        if not s:
            return 0
        try:
            return float(str(s).replace(".", "").replace(",", "."))
        except:
            return 0

    total_dealer = 0
    total_excel = 0
    for item in tabela:
        vt = to_float(item.get("valor_total", ""))
        sd = to_float(item.get("saldo", ""))
        baixa_dealer = float(item.get("valor_baixa_dealer", 0))
        baixa_excel = float(item.get("valor_baixa_excel", 0))
        diferenca = round(baixa_dealer - baixa_excel, 2)
        total_dealer += baixa_dealer
        total_excel += baixa_excel
        ws.append([
            item.get("cnpj", ""),
            item.get("nf_original", item.get("nf", "")),
            vt,
            sd,
            baixa_dealer,
            baixa_excel,
            diferenca,
        ])

    # Linha de totais
    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=1, value="TOTAIS").font = Font(bold=True, size=12)
    ws.cell(row=total_row, column=5, value=total_dealer).font = Font(bold=True, size=12)
    ws.cell(row=total_row, column=6, value=total_excel).font = Font(bold=True, size=12)
    ws.cell(row=total_row, column=7, value=round(total_dealer - total_excel, 2)).font = Font(bold=True, size=12)

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=7):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 18
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    loja = estado.get("loja_selecionada", "analise")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"analise_{loja}_{timestamp}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/api/iniciar", methods=["POST"])
def api_iniciar():
    estado["inicio_confirmado"] = True
    return jsonify({"ok": True})


@app.route("/api/pausar", methods=["POST"])
def api_pausar():
    if automacao:
        automacao.pausado = True
    if automacao_gaulesa:
        automacao_gaulesa.pausado = True
    if automacao_cancel:
        automacao_cancel.pausado = True
    return jsonify({"ok": True})


@app.route("/api/recomecar", methods=["POST"])
def api_recomecar():
    if automacao:
        automacao.pausado = False
    if automacao_gaulesa:
        automacao_gaulesa.pausado = False
    if automacao_cancel:
        automacao_cancel.pausado = False
    return jsonify({"ok": True})


@app.route("/api/exportar")
def api_exportar():
    """Gera e retorna um Excel com os resultados da rodagem."""
    tabela = estado.get("tabela_nfs", [])
    if not tabela:
        return jsonify({"ok": False, "erro": "Nenhum dado para exportar"}), 400

    STATUS_LABELS = {
        "sucesso": "Baixada",
        "pago": "Ja estava paga",
        "nao_encontrada": "Nao encontrada",
        "erro": "Erro",
        "processando": "Processando",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Baixas"

    # Cabeçalho
    headers = ["CNPJ", "NF", "Total da Baixa", "Valor Total da Nota", "Status"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dados
    for item in tabela:
        status_label = STATUS_LABELS.get(item.get("status", ""), item.get("status", ""))
        # Converte valor total (string BR "1.234,56" → float)
        valor_total_str = item.get("valor_total_nota", "")
        valor_total_num = 0
        if valor_total_str:
            try:
                valor_total_num = float(str(valor_total_str).replace(".", "").replace(",", "."))
            except:
                valor_total_num = 0
        ws.append([
            item.get("cnpj", ""),
            item.get("nf_original", item.get("nf", "")),
            float(item.get("valor", 0)),
            valor_total_num,
            status_label,
        ])

    # Formatação da colunas de valor (C e D)
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    # Ajusta largura das colunas
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # Salva em memória
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    loja = estado.get("loja_selecionada", "baixa")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"relatorio_{loja}_{timestamp}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/api/parar", methods=["POST"])
def api_parar():
    global estado
    estado["rodando"] = False
    if automacao:
        automacao.parar = True
        automacao.pausado = False
    if automacao_gaulesa:
        automacao_gaulesa.parar = True
        automacao_gaulesa.pausado = False
    if automacao_cancel:
        automacao_cancel.parar = True
        automacao_cancel.pausado = False
    return jsonify({"ok": True})


@app.route("/api/configurar_gaulesa", methods=["POST"])
def api_configurar_gaulesa():
    global automacao_gaulesa, estado
    arquivo_upload = request.files.get("arquivo") if request.content_type and "multipart" in request.content_type else None

    if not arquivo_upload or not arquivo_upload.filename:
        return jsonify({"ok": False, "erro": "Selecione o arquivo Excel da Gaulesa!"})

    if not arquivo_upload.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"ok": False, "erro": "Arquivo deve ser .xlsx ou .xls"})

    uploads_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
    os.makedirs(uploads_dir, exist_ok=True)
    caminho = os.path.join(uploads_dir, "gaulesa.xlsx")
    arquivo_upload.save(caminho)

    estado["loja_selecionada"] = "gaulesa"
    estado["log_mensagens"] = []
    estado["rodando"] = False
    estado["dealer_pronto"] = False
    estado["inicio_confirmado"] = False
    estado["browser_aberto"] = False
    estado["tabela_nfs"] = []
    estado["tabela_analise"] = []
    estado["nf_atual"] = ""

    automacao_gaulesa = AutomacaoGaulesa(caminho, estado)
    total = automacao_gaulesa.carregar_notas()
    estado["browser_aberto"] = True
    return jsonify({"ok": True, "total": total})


@app.route("/api/configurar_cancelamento", methods=["POST"])
def api_configurar_cancelamento():
    global automacao_cancel, estado
    arquivo_upload = request.files.get("arquivo") if request.content_type and "multipart" in request.content_type else None
    if not arquivo_upload or not arquivo_upload.filename:
        return jsonify({"ok": False, "erro": "Selecione o arquivo Excel!"})
    uploads_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
    os.makedirs(uploads_dir, exist_ok=True)
    caminho = os.path.join(uploads_dir, "cancelamento_gaulesa.xlsx")
    arquivo_upload.save(caminho)
    estado["loja_selecionada"] = "cancelamento_gaulesa"
    estado["log_mensagens"] = []
    estado["rodando"] = False
    estado["dealer_pronto"] = False
    estado["inicio_confirmado"] = False
    estado["tabela_nfs"] = []
    estado["nf_atual"] = ""
    automacao_cancel = AutomacaoCancelamento(caminho, estado)
    total = automacao_cancel.carregar_notas()
    return jsonify({"ok": True, "total": total})


@app.route("/api/comecar_cancelamento", methods=["POST"])
def api_comecar_cancelamento():
    global automacao_cancel, estado
    if automacao_cancel is None:
        return jsonify({"ok": False, "erro": "Configure o cancelamento primeiro!"})
    if estado["rodando"]:
        return jsonify({"ok": False, "erro": "Ja esta rodando!"})
    total = len(automacao_cancel.notas)
    estado["progresso"] = {"total": total, "processadas": 0, "sucesso": 0, "pago": 0, "nao_encontrada": 0, "baixada_anteriormente": 0, "erro": 0}
    estado["tabela_nfs"] = []
    estado["log_mensagens"] = []
    estado["rodando"] = True
    estado["dealer_pronto"] = False
    estado["inicio_confirmado"] = False
    thread = threading.Thread(target=automacao_cancel.executar_cancelamento, daemon=True)
    thread.start()
    return jsonify({"ok": True, "total": total})


@app.route("/api/comecar_gaulesa", methods=["POST"])
def api_comecar_gaulesa():
    global automacao_gaulesa, estado
    if automacao_gaulesa is None:
        return jsonify({"ok": False, "erro": "Configure a Gaulesa primeiro!"})

    if estado["rodando"]:
        return jsonify({"ok": False, "erro": "Ja esta rodando!"})

    total = len(automacao_gaulesa.notas)
    estado["progresso"] = {"total": total, "processadas": 0, "sucesso": 0, "pago": 0, "nao_encontrada": 0, "baixada_anteriormente": 0, "erro": 0}
    estado["tabela_nfs"] = []
    estado["log_mensagens"] = []
    estado["rodando"] = True
    estado["dealer_pronto"] = False
    estado["inicio_confirmado"] = False

    thread = threading.Thread(target=automacao_gaulesa.executar_tudo, daemon=True)
    thread.start()
    return jsonify({"ok": True, "total": total})


if __name__ == "__main__":
    print("=" * 50)
    print("  PAINEL DE CONTROLE - BAIXA DE NF")
    print("  Acesse: http://localhost:5000")
    print("=" * 50)
    app.run(host="127.0.0.1", port=5000, debug=False)
